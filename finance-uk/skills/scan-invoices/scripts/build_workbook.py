#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LOW_FILL = PatternFill("solid", fgColor="F4CCCC")
MEDIUM_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="D9EAD3")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the purchase ledger workbook from JSON outputs.")
    parser.add_argument("--validated-invoices", required=True, help="Path to validated_invoices.json")
    parser.add_argument("--validation-flags", required=True, help="Path to validation_flags.json")
    parser.add_argument("--processing-log", required=True, help="Path to processing_log.json")
    parser.add_argument("--output-path", required=True, help="Destination .xlsx path")
    return parser.parse_args()


def load_json(path: str) -> object:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def style_headers(ws, row: int) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_fit_columns(ws, min_width: int = 10, max_width: int = 42) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, min_width), max_width)


def build_invoice_register(ws, invoices: list[dict[str, object]]) -> None:
    headers = [
        "#",
        "Invoice Number",
        "Supplier",
        "Supplier VAT Reg",
        "Tax Point",
        "Net (£)",
        "VAT (£)",
        "Gross (£)",
        "VAT Rate",
        "Credit Note",
        "Quarter",
        "Confidence",
        "Source",
        "Notes",
        "Month Key",
    ]
    ws.title = "Invoice Register"
    ws.append(headers)
    style_headers(ws, 1)

    for index, invoice in enumerate(invoices, start=1):
        tax_point_iso = invoice.get("tax_point_iso") or ""
        tax_point_value = invoice.get("tax_point_date") or ""
        if tax_point_iso:
            try:
                tax_point_value = datetime.strptime(str(tax_point_iso), "%Y-%m-%d").date()
            except ValueError:
                pass

        month_key = ""
        if tax_point_iso:
            try:
                month_key = datetime.strptime(str(tax_point_iso), "%Y-%m-%d").strftime("%b %Y")
            except ValueError:
                month_key = ""

        ws.append(
            [
                index,
                invoice.get("invoice_number"),
                invoice.get("supplier_name"),
                invoice.get("supplier_vat_reg"),
                tax_point_value,
                invoice.get("net_total"),
                invoice.get("vat_amount"),
                invoice.get("gross_total"),
                (invoice.get("vat_rate_percent") or 0) / 100 if invoice.get("vat_rate_percent") is not None else None,
                "Yes" if invoice.get("is_credit_note") else "No",
                invoice.get("quarter_label"),
                str(invoice.get("confidence", "")).title(),
                invoice.get("page_ref"),
                invoice.get("notes"),
                month_key,
            ]
        )

    last_data_row = max(ws.max_row, 2)
    totals_row = last_data_row + 1
    ws[f"E{totals_row}"] = "Totals"
    for column_letter in ("F", "G", "H"):
        ws[f"{column_letter}{totals_row}"] = f"=SUM({column_letter}2:{column_letter}{last_data_row})"
        ws[f"{column_letter}{totals_row}"].fill = TOTAL_FILL
        ws[f"{column_letter}{totals_row}"].font = Font(bold=True)
    ws[f"E{totals_row}"].fill = TOTAL_FILL
    ws[f"E{totals_row}"].font = Font(bold=True)

    for row in range(2, totals_row + 1):
        ws[f"F{row}"].number_format = '£#,##0.00'
        ws[f"G{row}"].number_format = '£#,##0.00'
        ws[f"H{row}"].number_format = '£#,##0.00'
        ws[f"I{row}"].number_format = "0.0%"
        ws[f"E{row}"].number_format = "DD/MM/YYYY"

    if last_data_row >= 2:
        data_range = f"A2:O{last_data_row}"
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=['$L2="Low"'], fill=LOW_FILL),
        )
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=['$L2="Medium"'], fill=MEDIUM_FILL),
        )

    ws.auto_filter.ref = f"A1:O{totals_row}"
    ws.freeze_panes = "A2"
    ws.column_dimensions["N"].width = 42
    ws.column_dimensions["O"].hidden = True
    auto_fit_columns(ws)


def build_vat_summary(ws, invoices: list[dict[str, object]]) -> None:
    suppliers = sorted({str(invoice.get("supplier_name") or "") for invoice in invoices if invoice.get("supplier_name")})
    month_pairs = sorted(
        {
            (
                datetime.strptime(str(invoice["tax_point_iso"]), "%Y-%m-%d").replace(day=1),
                datetime.strptime(str(invoice["tax_point_iso"]), "%Y-%m-%d").strftime("%b %Y"),
            )
            for invoice in invoices
            if invoice.get("tax_point_iso")
        },
        key=lambda item: item[0],
    )
    months = [label for _, label in month_pairs]

    quarter_pairs = sorted(
        {
            (
                (dt.year if dt.month >= 4 else dt.year - 1),
                (1 if 4 <= dt.month <= 6 else 2 if 7 <= dt.month <= 9 else 3 if 10 <= dt.month <= 12 else 4),
                str(invoice.get("quarter_label") or ""),
            )
            for invoice in invoices
            if invoice.get("tax_point_iso")
            for dt in [datetime.strptime(str(invoice["tax_point_iso"]), "%Y-%m-%d")]
            if invoice.get("quarter_label")
        },
        key=lambda item: (item[0], item[1]),
    )
    quarters = [label for _, _, label in quarter_pairs]

    ws.title = "VAT Summary"
    ws["A1"] = "By supplier"
    ws["A2"] = "Supplier"
    ws["B2"] = "Invoice Count"
    ws["C2"] = "Net (£)"
    ws["D2"] = "VAT (£)"
    style_headers(ws, 2)

    supplier_start = 3
    for offset, supplier in enumerate(suppliers, start=supplier_start):
        ws[f"A{offset}"] = supplier
        ws[f"B{offset}"] = f'=COUNTIF(\'Invoice Register\'!$C:$C, A{offset})'
        ws[f"C{offset}"] = f'=SUMIF(\'Invoice Register\'!$C:$C, A{offset}, \'Invoice Register\'!$F:$F)'
        ws[f"D{offset}"] = f'=SUMIF(\'Invoice Register\'!$C:$C, A{offset}, \'Invoice Register\'!$G:$G)'

    month_header_row = max(supplier_start + len(suppliers) + 2, 6)
    ws[f"A{month_header_row}"] = "By month"
    ws[f"A{month_header_row + 1}"] = "Month"
    ws[f"B{month_header_row + 1}"] = "Invoice Count"
    ws[f"C{month_header_row + 1}"] = "Net (£)"
    ws[f"D{month_header_row + 1}"] = "VAT (£)"
    style_headers(ws, month_header_row + 1)

    month_start = month_header_row + 2
    for offset, month in enumerate(months, start=month_start):
        ws[f"A{offset}"] = month
        ws[f"B{offset}"] = f'=COUNTIF(\'Invoice Register\'!$O:$O, A{offset})'
        ws[f"C{offset}"] = f'=SUMIF(\'Invoice Register\'!$O:$O, A{offset}, \'Invoice Register\'!$F:$F)'
        ws[f"D{offset}"] = f'=SUMIF(\'Invoice Register\'!$O:$O, A{offset}, \'Invoice Register\'!$G:$G)'

    quarter_header_row = max(month_start + len(months) + 2, month_header_row + 6)
    ws[f"A{quarter_header_row}"] = "By quarter"
    ws[f"A{quarter_header_row + 1}"] = "Quarter"
    ws[f"B{quarter_header_row + 1}"] = "Invoice Count"
    ws[f"C{quarter_header_row + 1}"] = "Net (£)"
    ws[f"D{quarter_header_row + 1}"] = "VAT (£)"
    style_headers(ws, quarter_header_row + 1)

    quarter_start = quarter_header_row + 2
    for offset, quarter in enumerate(quarters, start=quarter_start):
        ws[f"A{offset}"] = quarter
        ws[f"B{offset}"] = f'=COUNTIF(\'Invoice Register\'!$K:$K, A{offset})'
        ws[f"C{offset}"] = f'=SUMIF(\'Invoice Register\'!$K:$K, A{offset}, \'Invoice Register\'!$F:$F)'
        ws[f"D{offset}"] = f'=SUMIF(\'Invoice Register\'!$K:$K, A{offset}, \'Invoice Register\'!$G:$G)'

    for row in range(3, ws.max_row + 1):
        ws[f"C{row}"].number_format = '£#,##0.00'
        ws[f"D{row}"].number_format = '£#,##0.00'

    auto_fit_columns(ws)
    ws.freeze_panes = "A2"


def build_validation_report(ws, flags: list[dict[str, object]]) -> None:
    ws.title = "Validation Report"
    ws.append(["Invoice Number", "Supplier", "Flag Type", "Description", "Expected Value", "Actual Value"])
    style_headers(ws, 1)

    for flag in flags:
        ws.append(
            [
                flag.get("invoice_number"),
                flag.get("supplier"),
                flag.get("flag_type"),
                flag.get("description"),
                flag.get("expected_value"),
                flag.get("actual_value"),
            ]
        )

    for row in range(2, ws.max_row + 1):
        if ws[f"C{row}"].value == "ERROR":
            for cell in ws[row]:
                cell.fill = LOW_FILL
        elif ws[f"C{row}"].value == "WARNING":
            for cell in ws[row]:
                cell.fill = MEDIUM_FILL

    ws.auto_filter.ref = f"A1:F{max(ws.max_row, 2)}"
    ws.freeze_panes = "A2"
    auto_fit_columns(ws)


def build_processing_log(ws, processing_log: list[dict[str, object]]) -> None:
    invoice_pages = sum(1 for item in processing_log if item.get("classification") == "Invoice")
    credit_note_pages = sum(1 for item in processing_log if item.get("classification") == "Credit Note")
    skipped_pages = sum(1 for item in processing_log if item.get("extraction_status") == "Skipped")
    errored_pages = sum(1 for item in processing_log if item.get("extraction_status") == "Error")

    ws.title = "Processing Log"
    ws["A1"] = (
        f"{invoice_pages} invoice page(s), {credit_note_pages} credit note page(s), "
        f"{skipped_pages} skipped page(s), {errored_pages} error page(s)"
    )
    ws.append([])
    ws.append(["PDF Filename", "Page Number", "Classification", "Invoice Number", "Extraction Status"])
    style_headers(ws, 3)

    for item in processing_log:
        ws.append(
            [
                item.get("pdf_filename"),
                item.get("page_number"),
                item.get("classification"),
                item.get("invoice_number"),
                item.get("extraction_status"),
            ]
        )

    ws.auto_filter.ref = f"A3:E{max(ws.max_row, 3)}"
    ws.freeze_panes = "A4"
    auto_fit_columns(ws)


def main() -> int:
    args = parse_args()
    invoices = load_json(args.validated_invoices)
    flags = load_json(args.validation_flags)
    processing_log = load_json(args.processing_log)

    if not isinstance(invoices, list) or not isinstance(flags, list) or not isinstance(processing_log, list):
        raise SystemExit("Expected JSON arrays for invoices, flags, and processing log inputs.")

    output_path = Path(args.output_path).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    build_invoice_register(wb.active, invoices)
    build_vat_summary(wb.create_sheet(), invoices)
    build_validation_report(wb.create_sheet(), flags)
    build_processing_log(wb.create_sheet(), processing_log)
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    wb.save(output_path)

    print(f"Workbook saved to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
